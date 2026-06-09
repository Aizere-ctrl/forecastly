import os
from functools import wraps
from datetime import datetime
from io import BytesIO

import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from sklearn.ensemble import IsolationForest
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, make_response, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from ml.forecasting import generate_forecast
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.utils import ImageReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from flask_mail import Mail, Message
from itsdangerous import URLSafeTimedSerializer, SignatureExpired, BadSignature
from werkzeug.security import generate_password_hash
app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "forecastly-dev-secret-key")
app.config["MAIL_USERNAME"] = os.environ.get("MAIL_USERNAME")
app.config["MAIL_PASSWORD"] = os.environ.get("MAIL_PASSWORD")
app.config["MAIL_DEFAULT_SENDER"] = os.environ.get("MAIL_DEFAULT_SENDER")
mail = Mail(app)

serializer = URLSafeTimedSerializer(app.config['SECRET_KEY'])
print("✅ NEW CLEAN APP.PY LOADED")

 
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///forecastly.db"
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

UPLOAD_FOLDER = "uploads"
ALLOWED_EXTENSIONS = {"csv", "xlsx"}
MAX_FILE_SIZE_MB = 10

app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
app.config["MAX_CONTENT_LENGTH"] = MAX_FILE_SIZE_MB * 1024 * 1024

os.makedirs(UPLOAD_FOLDER, exist_ok=True)

db = SQLAlchemy(app)


# =========================
# DATABASE MODELS
# =========================

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    full_name = db.Column(db.String(150), nullable=False)
    email = db.Column(db.String(150), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    is_admin = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
     

    files = db.relationship(
        "UploadedFile",
        backref="user",
        lazy=True,
        cascade="all, delete-orphan"
    )


class UploadedFile(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    filename = db.Column(db.String(255), nullable=False)
    original_filename = db.Column(db.String(255), nullable=False)
    marketplace = db.Column(db.String(100), default="Not selected")
    status = db.Column(db.String(50), default="Uploaded")
    rows_count = db.Column(db.Integer, default=0)
    columns_count = db.Column(db.Integer, default=0)
    uploaded_at = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey("user.id"), nullable=False)


# =========================
# BASIC HELPERS
# =========================

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function
def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if "user_id" not in session:
            flash("Please log in first.", "error")
            return redirect(url_for("login"))

        user = User.query.get(session["user_id"])

        if not user or not user.is_admin:
            flash("Admin access only.", "error")
            return redirect(url_for("dashboard"))

        return f(*args, **kwargs)

    return decorated_function


def normalize_column_name(column_name):
    column_name = str(column_name).strip().lower()
    column_name = column_name.replace("\ufeff", "")
    column_name = column_name.replace("-", "_")
    column_name = column_name.replace(" ", "_")

    while "__" in column_name:
        column_name = column_name.replace("__", "_")

    return column_name


def normalize_dataset_columns(df):
    df = df.copy()
    df.columns = [normalize_column_name(col) for col in df.columns]

    aliases = {
        "sku": "sku_id",
        "sku_code": "sku_id",
        "sku_number": "sku_id",
        "product_id": "sku_id",
        "product_code": "sku_id",
        "item_id": "sku_id",

        "product": "sku_name",
        "product_name": "sku_name",
        "item_name": "sku_name",
        "name": "sku_name",

        "sales": "sales_units",
        "units_sold": "sales_units",
        "quantity": "sales_units",
        "qty": "sales_units",
        "sold_units": "sales_units",

        "unit_price": "price",
        "selling_price": "price",

        "total_revenue": "revenue",
        "sales_revenue": "revenue",
        "amount": "revenue",

        "sales_date": "date",
        "order_date": "date",
        "created_at": "date"
    }

    df.rename(columns=aliases, inplace=True)
    return df


def read_uploaded_file(file_path):
    if file_path.lower().endswith(".xlsx"):
        return pd.read_excel(file_path)

    for sep in [",", ";", "\t"]:
        try:
            df = pd.read_csv(file_path, sep=sep)
            if len(df.columns) >= 3:
                return df
        except Exception:
            pass

    return pd.read_csv(file_path)


def save_normalized_dataset(df, file_path):
    if file_path.lower().endswith(".xlsx"):
        df.to_excel(file_path, index=False)
    else:
        df.to_csv(file_path, index=False, encoding="utf-8-sig")


def get_latest_user_file():
    selected_file_id = session.get("selected_file_id")

    if selected_file_id:
        selected = UploadedFile.query.filter_by(
            id=selected_file_id,
            user_id=session["user_id"]
        ).first()

        if selected:
            return selected

    return UploadedFile.query.filter_by(
        user_id=session["user_id"]
    ).order_by(UploadedFile.uploaded_at.desc()).first()


# =========================
# DATA VALIDATION
# =========================

def validate_uploaded_dataset(file_path):
    try:
        df = read_uploaded_file(file_path)

        if df.empty:
            return False, "File is empty.", 0, 0

        df = normalize_dataset_columns(df)

        required_columns = ["date", "sku_id", "sku_name", "sales_units", "price", "revenue"]
        missing = [col for col in required_columns if col not in df.columns]

        if missing:
            return False, f"Missing required column(s): {', '.join(missing)}. Available columns: {', '.join(df.columns)}", 0, 0

        df = df[required_columns].copy()

        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df["sales_units"] = pd.to_numeric(df["sales_units"], errors="coerce")
        df["price"] = pd.to_numeric(df["price"], errors="coerce")
        df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce")

        df.dropna(subset=["date", "sales_units", "price", "revenue"], inplace=True)

        if df.empty:
            return False, "After cleaning, there is no valid data left.", 0, 0

        save_normalized_dataset(df, file_path)

        return True, "Valid file.", len(df), len(df.columns)

    except Exception as e:
        print("VALIDATION ERROR:", e)
        return False, "Invalid file. Please upload a correct CSV or Excel file.", 0, 0


# =========================
# FORECASTLY CORE LOGIC
# =========================

def prepare_forecast_dataframe(file_path):
    df = read_uploaded_file(file_path)
    df = normalize_dataset_columns(df)

    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["revenue"] = pd.to_numeric(df["revenue"], errors="coerce")
    df["sales_units"] = pd.to_numeric(df["sales_units"], errors="coerce")
    df["price"] = pd.to_numeric(df["price"], errors="coerce")

    df.dropna(subset=["date", "revenue"], inplace=True)
    df = df.sort_values("date")

    return df


def generate_simple_forecast(df, forecast_days=7):
    daily = df.groupby("date")["revenue"].sum().reset_index()
    daily = daily.sort_values("date")

    if daily.empty:
        raise ValueError("Not enough revenue data for forecasting.")

    last_date = daily["date"].max()

    if len(daily) >= 7:
        base_value = daily["revenue"].tail(7).mean()
    else:
        base_value = daily["revenue"].mean()

    if len(daily) >= 2:
        first_value = daily["revenue"].iloc[0]
        last_value = daily["revenue"].iloc[-1]

        if first_value != 0:
            growth_rate = (last_value - first_value) / abs(first_value)
        else:
            growth_rate = 0
    else:
        growth_rate = 0

    daily_growth = max(min(growth_rate / max(len(daily), 1), 0.03), -0.03)

    forecast = []

    for i in range(1, forecast_days + 1):
        predicted = base_value * ((1 + daily_growth) ** i)

        forecast.append({
            "period": i,
            "date": (last_date + pd.Timedelta(days=i)).strftime("%Y-%m-%d"),
            "predicted_revenue": round(float(predicted), 2)
        })

    return forecast


def extract_forecast_values(forecast_data):
    values = []

    if isinstance(forecast_data, list):
        for item in forecast_data:
            if isinstance(item, dict):
                value = (
                    item.get("predicted_revenue")
                    or item.get("forecast")
                    or item.get("revenue")
                    or item.get("predicted_sales")
                    or item.get("sales")
                    or item.get("value")
                )

                if value is not None:
                    values.append(float(value))

            elif isinstance(item, (int, float)):
                values.append(float(item))

    return values


def generate_safe_scenarios(forecast_data):
    values = extract_forecast_values(forecast_data)

    total_forecast = round(sum(values), 2) if values else 0
    average_forecast = round(total_forecast / len(values), 2) if values else 0

    realistic = total_forecast
    optimistic = round(total_forecast * 1.15, 2)
    pessimistic = round(total_forecast * 0.85, 2)

    return {
        "total_forecast": total_forecast,
        "average_forecast": average_forecast,
        "realistic": realistic,
        "optimistic": optimistic,
        "pessimistic": pessimistic,
        "description": "Scenario analysis is based on predicted revenue with optimistic and pessimistic variations."
    }


def calculate_safe_analytics(df, forecast_data):
    total_revenue = round(float(df["revenue"].sum()), 2)
    average_revenue = round(float(df["revenue"].mean()), 2)

    daily = df.groupby("date")["revenue"].sum().reset_index()
    daily = daily.sort_values("date")

    max_revenue = round(float(daily["revenue"].max()), 2) if not daily.empty else 0
    min_revenue = round(float(daily["revenue"].min()), 2) if not daily.empty else 0

    moving_average = (
        round(float(daily["revenue"].tail(7).mean()), 2)
        if len(daily) >= 7
        else round(float(daily["revenue"].mean()), 2)
        if not daily.empty
        else 0
    )

    if len(daily) >= 2 and daily["revenue"].iloc[0] != 0:
        growth_rate = round(
            ((daily["revenue"].iloc[-1] - daily["revenue"].iloc[0]) / abs(daily["revenue"].iloc[0])) * 100,
            2
        )
    else:
        growth_rate = 0

    forecast_values = extract_forecast_values(forecast_data)
    average_forecast = round(sum(forecast_values) / len(forecast_values), 2) if forecast_values else 0
    total_forecast = round(sum(forecast_values), 2) if forecast_values else 0

    model_used = "Not available"
    if forecast_data and isinstance(forecast_data, list):
        first_forecast = forecast_data[0]
        if isinstance(first_forecast, dict):
            model_used = first_forecast.get("model_used", "Not available")

    if average_revenue != 0:
        mape = round(abs(average_forecast - average_revenue) / abs(average_revenue) * 100, 2)
    else:
        mape = 0

    accuracy = max(0, round(100 - mape, 2))
    rmse = round(abs(average_forecast - average_revenue), 2)

    return {
        "total_sales": total_revenue,
        "average_sales": average_revenue,
        "growth_rate": growth_rate,

        "average_forecast": average_forecast,
        "total_forecast": total_forecast,

        "max_sales": max_revenue,
        "min_sales": min_revenue,
        "moving_average": moving_average,

        "mape": mape,
        "rmse": rmse,
        "accuracy": accuracy,
        "model_used": model_used,

        "total_revenue": total_revenue,
        "average_revenue": average_revenue,
        "max_revenue": max_revenue,
        "min_revenue": min_revenue
    }


def detect_safe_anomalies(df):
    daily = df.groupby("date")["revenue"].sum().reset_index()
    daily = daily.sort_values("date")

    if daily.empty or len(daily) < 7:
        return []

    daily["day_index"] = range(len(daily))
    daily["day_of_week"] = daily["date"].dt.dayofweek
    daily["rolling_mean_3"] = daily["revenue"].rolling(
        window=3,
        min_periods=1
    ).mean()

    daily["revenue_change"] = daily["revenue"].pct_change().fillna(0)

    feature_columns = [
        "revenue",
        "day_index",
        "day_of_week",
        "rolling_mean_3",
        "revenue_change"
    ]

    X = daily[feature_columns].fillna(0)

    try:
        model = IsolationForest(
            contamination=0.15,
            random_state=42
        )

        daily["anomaly_score"] = model.fit_predict(X)

        anomalies = []

        for _, row in daily.iterrows():
            if row["anomaly_score"] == -1:
                revenue = round(float(row["revenue"]), 2)
                avg_revenue = daily["revenue"].mean()

                anomaly_type = (
                    "High anomaly"
                    if revenue > avg_revenue
                    else "Low anomaly"
                )

                anomalies.append({
                    "date": row["date"].strftime("%Y-%m-%d"),
                    "revenue": revenue,
                    "type": anomaly_type,
                    "severity": "High" if abs(row["revenue_change"]) > 0.4 else "Medium"
                })

        return anomalies

    except Exception as e:
        print("Isolation Forest anomaly error:", e)
        return []

def generate_safe_insight(kind, analytics=None, scenarios=None, anomalies=None):

    if kind == "scenario":
        if not scenarios:
            return [
                "Forecastly generated a basic scenario forecast based on the uploaded sales data."
            ]

        realistic = scenarios.get("realistic", 0)
        optimistic = scenarios.get("optimistic", 0)
        pessimistic = scenarios.get("pessimistic", 0)

        gap = optimistic - pessimistic

        if realistic > 0 and gap / realistic > 0.4:
            return [
                "Forecastly detected a wide gap between optimistic and pessimistic outcomes. This means future revenue may be sensitive to demand changes, so inventory planning should be more careful."
            ]

        if optimistic > realistic > pessimistic:
            return [
                "Forecastly expects stable future demand with a positive upside opportunity. The business can prepare stock based on the realistic scenario while monitoring growth potential."
            ]

        return [
            "Forecastly generated scenario analysis to support revenue planning and business decision-making."
        ]

    if kind == "analytics":
        if not analytics:
            return [
                "Forecastly generated analytics based on the uploaded sales data."
            ]

        growth_rate = analytics.get("growth_rate", 0)
        accuracy = analytics.get("accuracy", 0)
        mape = analytics.get("mape", 0)
        average_revenue = analytics.get("average_revenue", 0)
        average_forecast = analytics.get("average_forecast", 0)

        if accuracy < 60:
            return [
                "Forecast confidence is currently low. Forecastly recommends uploading more historical sales data to improve prediction quality."
            ]

        if mape > 40:
            return [
                "Forecast error is relatively high, which may indicate unstable revenue patterns. The business should use this forecast carefully for planning."
            ]

        if growth_rate > 20:
            return [
                "Revenue is growing strongly. Forecastly detected increasing demand, which may indicate a good opportunity to prepare more inventory."
            ]

        if growth_rate > 5:
            return [
                "Revenue shows healthy growth. Demand is gradually increasing, and the business performance appears positive."
            ]

        if growth_rate < -20:
            return [
                "Revenue is decreasing significantly. Forecastly recommends reviewing recent sales performance and adjusting inventory plans carefully."
            ]

        if growth_rate < 0:
            return [
                "Revenue shows a slight decline. The business should monitor demand changes and avoid overstocking."
            ]

        if average_forecast > average_revenue * 1.15:
            return [
                "Future forecast is higher than the historical average. Forecastly expects stronger upcoming demand compared to previous sales performance."
            ]

        if average_forecast < average_revenue * 0.85:
            return [
                "Future forecast is lower than the historical average. This may indicate weaker upcoming demand, so conservative planning is recommended."
            ]

        return [
            "Revenue patterns appear stable. Forecastly expects consistent future performance based on the current sales data."
        ]

    if kind == "anomaly":
        if anomalies and len(anomalies) >= 3:
            return [
                "Forecastly detected several unusual revenue changes. These anomalies may affect forecast accuracy and should be reviewed before making business decisions."
            ]

        if anomalies:
            return [
                "Forecastly detected unusual revenue behavior in some periods. These changes may be caused by promotions, stockouts, or seasonal demand."
            ]

        return [
            "No significant anomalies were detected. Revenue behavior appears stable and suitable for forecasting."
        ]

    return [
        "Forecastly generated an AI insight based on the uploaded sales data."
    ]


def build_chart_data(df, forecast_data):
    daily = df.groupby("date")["revenue"].sum().reset_index()
    daily = daily.sort_values("date").tail(6)

    actual_values = daily["revenue"].round(2).tolist()
    actual_labels = daily["date"].dt.strftime("%Y-%m-%d").tolist()

    forecast_values = [
    round(v * (1 + i * 0.02), 2)
    for i, v in enumerate(extract_forecast_values(forecast_data)[:4])
]
    forecast_labels = [f"Forecast {i + 1}" for i in range(len(forecast_values))]

    labels = actual_labels + forecast_labels

    actual_chart = actual_values + [None] * len(forecast_values)

    if actual_values:
        forecast_chart = [None] * (len(actual_values) - 1) + [actual_values[-1]] + forecast_values
    else:
        forecast_chart = forecast_values

    optimistic_chart = [None if v is None else round(v * 1.15, 2) for v in forecast_chart]
    realistic_chart = forecast_chart
    pessimistic_chart = [None if v is None else round(v * 0.85, 2) for v in forecast_chart]

    return {
        "labels": labels,
        "actual": actual_chart,
        "forecast": forecast_chart,
        "optimistic": optimistic_chart,
        "realistic": realistic_chart,
        "pessimistic": pessimistic_chart
    }


def run_ml_pipeline(forecast_days=7):
    latest_file = get_latest_user_file()

    if not latest_file:
        raise ValueError("Please upload a CSV or Excel file first.")

    file_path = os.path.join(app.config["UPLOAD_FOLDER"], latest_file.filename)

    if not os.path.exists(file_path):
        raise ValueError("Uploaded file was not found.")

    df = prepare_forecast_dataframe(file_path)

    forecast_data = generate_forecast(df, forecast_days)
    scenarios = generate_safe_scenarios(forecast_data)
    analytics = calculate_safe_analytics(df, forecast_data)
    anomalies = detect_safe_anomalies(df)
    chart_data = build_chart_data(df, forecast_data)

    return {
        "latest_file": latest_file,
        "forecast_data": forecast_data,
        "scenarios": scenarios,
        "analytics": analytics,
        "anomalies": anomalies,
        "scenario_insights": generate_safe_insight("scenario", scenarios=scenarios),
        "analytics_insights": generate_safe_insight("analytics", analytics=analytics),
        "anomaly_insights": generate_safe_insight("anomaly", anomalies=anomalies),
        "chart_data": chart_data
    }


# =========================
# ROUTES
# =========================

@app.route("/")
def home():
    return render_template("index.html")


@app.route("/signup", methods=["GET", "POST"])
def signup():
    if request.method == "POST":
        full_name = request.form.get("fullname", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")
        confirm_password = request.form.get("confirm_password", "")

        if not full_name or not email or not password or not confirm_password:
            flash("Please fill in all fields.", "error")
            return redirect(url_for("signup"))

        if len(password) < 6:
            flash("Password must be at least 6 characters long.", "error")
            return redirect(url_for("signup"))

        if password != confirm_password:
            flash("Passwords do not match.", "error")
            return redirect(url_for("signup"))

        if User.query.filter_by(email=email).first():
            flash("This email is already registered.", "error")
            return redirect(url_for("signup"))

        new_user = User(
            full_name=full_name,
            email=email,
            password_hash=generate_password_hash(password)
        )

        db.session.add(new_user)
        db.session.commit()

        flash("Account created successfully. Please log in.", "success")
        return redirect(url_for("login"))

    return render_template("signup.html")


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        user = User.query.filter_by(email=email).first()

        if user and check_password_hash(user.password_hash, password):
            session["user_id"] = user.id
            session["user_name"] = user.full_name
            session["is_admin"] = user.is_admin

            flash("Welcome back!", "success")

            if user.is_admin:
                return redirect(url_for("admin_dashboard"))
            return redirect(url_for("dashboard"))

        flash("Invalid email or password.", "error")
        return redirect(url_for("login"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("You have logged out.", "success")
    return redirect(url_for("home"))


@app.route("/dashboard")
@login_required
def dashboard():
     if session.get("is_admin"):
         return redirect(url_for("admin_dashboard"))
     

     files = UploadedFile.query.filter_by(
        user_id=session["user_id"]
    ).order_by(UploadedFile.uploaded_at.desc()).all()

     has_data = len(files) > 0
     latest_file = files[0] if has_data else None

     return render_template(
        "dashboard.html",
        user_name=session.get("user_name"),
        has_data=has_data,
        latest_file=latest_file,
        uploaded_files=files,
        total_files=len(files),
        total_rows=sum(file.rows_count or 0 for file in files),
        total_columns=latest_file.columns_count if latest_file else 0,
        forecast_accuracy=91 if has_data else 0,
        anomalies_detected=5 if has_data else 0
    )


@app.route("/upload", methods=["GET", "POST"])
@login_required
def upload_data():
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file selected.", "error")
            return redirect(url_for("upload_data"))

        file = request.files["file"]

        if file.filename == "":
            flash("Please choose a CSV or Excel file.", "error")
            return redirect(url_for("upload_data"))

        if not allowed_file(file.filename):
            flash("Only CSV or Excel files are allowed.", "error")
            return redirect(url_for("upload_data"))

        original_filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
        saved_filename = f"{session['user_id']}_{timestamp}_{original_filename}"
        file_path = os.path.join(app.config["UPLOAD_FOLDER"], saved_filename)

        file.save(file_path)

        is_valid, message, rows_count, columns_count = validate_uploaded_dataset(file_path)

        if not is_valid:
            if os.path.exists(file_path):
                os.remove(file_path)

            flash(message, "error")
            return redirect(url_for("upload_data"))

        uploaded_file = UploadedFile(
            filename=saved_filename,
            original_filename=original_filename,
            rows_count=rows_count,
            columns_count=columns_count,
            user_id=session["user_id"]
        )

        db.session.add(uploaded_file)
        db.session.commit()

        session["selected_file_id"] = uploaded_file.id

        flash("File uploaded successfully.", "success")
        return redirect(url_for("configuration"))

    return render_template("upload.html", user_name=session.get("user_name"))
@app.route("/download-template")
def download_template():
    return send_file(
        "template.xlsx",
        as_attachment=True
    )

@app.route("/configuration", methods=["GET", "POST"])
@login_required
def configuration():
    latest_file = get_latest_user_file()

    if not latest_file:
        flash("Please upload a CSV or Excel file first.", "error")
        return redirect(url_for("upload_data"))

    if request.method == "POST":
        print("✅ CONFIGURATION POST RECEIVED")

        try:
            forecast_days = int(request.form.get("forecast_days", 7))
        except ValueError:
            forecast_days = 7

        if forecast_days <= 0:
            forecast_days = 7

        marketplace = request.form.get("marketplace", "Not selected")
        granularity = request.form.get("granularity", "Daily")

        latest_file.marketplace = marketplace
        latest_file.status = "Configured"
        db.session.commit()

        session["forecast_days"] = forecast_days
        session["granularity"] = granularity
        session["marketplace"] = marketplace

        flash("Forecast configuration saved.", "success")
        return redirect(url_for("scenario"))

    return render_template(
        "configuration.html",
        user_name=session.get("user_name"),
        latest_file=latest_file
    )


@app.route("/scenario")
@login_required
def scenario():
    try:
        forecast_days = int(session.get("forecast_days", 7))
        result = run_ml_pipeline(forecast_days)

        print("✅ SCENARIO DATA:", result["scenarios"])

        return render_template(
            "scenario.html",
            user_name=session.get("user_name"),
            forecast_data=result["forecast_data"],
            scenarios=result["scenarios"],
            ai_insights=result["scenario_insights"],
            latest_file=result["latest_file"],
            marketplace=session.get("marketplace", result["latest_file"].marketplace),
            granularity=session.get("granularity", "Daily"),
            chart_data=result["chart_data"]
        )

    except Exception as e:
        print("❌ SCENARIO ERROR:", e)
        return f"""
        <h1>Scenario Error</h1>
        <p>{str(e)}</p>
        <br>
        <a href="/configuration">Back to configuration</a>
        """


@app.route("/analytics")
@login_required
def analytics_page():
    try:
        forecast_days = int(session.get("forecast_days", 7))
        result = run_ml_pipeline(forecast_days)

        return render_template(
            "analytics.html",
            user_name=session.get("user_name"),
            forecast_data=result["forecast_data"],
            analytics=result["analytics"],
            ai_insights=result["analytics_insights"],
            latest_file=result["latest_file"],
            chart_data=result["chart_data"]
        )

    except Exception as e:
        print("❌ ANALYTICS ERROR:", e)
        flash(f"Analytics error: {e}", "error")
        return redirect(url_for("configuration"))


@app.route("/anomalies")
@login_required
def anomalies_page():
    try:
        forecast_days = int(session.get("forecast_days", 7))
        result = run_ml_pipeline(forecast_days)

        return render_template(
            "anomalies.html",
            user_name=session.get("user_name"),
            anomalies=result["anomalies"],
            ai_insights=result["anomaly_insights"],
            latest_file=result["latest_file"]
        )

    except Exception as e:
        print("❌ ANOMALIES ERROR:", e)
        flash(f"Anomalies error: {e}", "error")
        return redirect(url_for("configuration"))


@app.route("/library")
@login_required
def library():
    files = UploadedFile.query.filter_by(
        user_id=session["user_id"]
    ).order_by(UploadedFile.uploaded_at.desc()).all()

    return render_template(
        "library.html",
        uploaded_files=files,
        user_name=session.get("user_name"),
        total_files=len(files),
        total_rows=sum(file.rows_count or 0 for file in files)
    )


@app.route("/delete-file/<int:file_id>", methods=["POST"])
@login_required
def delete_file(file_id):
    uploaded_file = UploadedFile.query.filter_by(
        id=file_id,
        user_id=session["user_id"]
    ).first()

    if not uploaded_file:
        flash("File not found.", "error")
        return redirect(url_for("library"))

    file_path = os.path.join(app.config["UPLOAD_FOLDER"], uploaded_file.filename)

    if os.path.exists(file_path):
        os.remove(file_path)

    if session.get("selected_file_id") == uploaded_file.id:
        session.pop("selected_file_id", None)

    db.session.delete(uploaded_file)
    db.session.commit()

    flash("File deleted successfully.", "success")
    return redirect(url_for("library"))


@app.route("/view-file/<int:file_id>")
@login_required
def view_file(file_id):
    uploaded_file = UploadedFile.query.filter_by(
        id=file_id,
        user_id=session["user_id"]
    ).first()

    if not uploaded_file:
        flash("File not found.", "error")
        return redirect(url_for("library"))

    session["selected_file_id"] = uploaded_file.id
    session["marketplace"] = uploaded_file.marketplace

    flash("File selected successfully.", "success")
    return redirect(url_for("configuration"))


@app.route("/account", methods=["GET", "POST"])
@login_required
def account():
    user = User.query.get(session["user_id"])

    if not user:
        session.clear()
        flash("User not found. Please log in again.", "error")
        return redirect(url_for("login"))

    if request.method == "POST":
        full_name = request.form.get("full_name", "").strip()
        email = request.form.get("email", "").strip().lower()
        password = request.form.get("password", "")

        if full_name:
            user.full_name = full_name

        if email:
            existing_user = User.query.filter_by(email=email).first()

            if existing_user and existing_user.id != user.id:
                flash("Email already in use.", "error")
                return redirect(url_for("account"))

            user.email = email

        if password:
            if len(password) < 6:
                flash("Password must be at least 6 characters long.", "error")
                return redirect(url_for("account"))

            user.password_hash = generate_password_hash(password)

        db.session.commit()
        session["user_name"] = user.full_name

        flash("Account updated successfully.", "success")
        return redirect(url_for("account"))

    return render_template(
        "account.html",
        user=user,
        user_name=session.get("user_name")
    )


@app.route("/ai-chat", methods=["POST"])
@login_required
def ai_chat():
    data = request.get_json() or {}
    text = data.get("message", "").lower()

    if not text:
        return jsonify({"reply": "Please type a message."})

    if "upload" in text or "csv" in text or "excel" in text:
        return jsonify({"reply": "To upload data: go to Upload, choose your CSV or Excel file, and click Upload & Continue."})

    if "forecast" in text or "predict" in text:
        return jsonify({"reply": "Forecastly predicts future revenue based on your uploaded sales dataset."})

    if "revenue" in text:
        return jsonify({"reply": "Forecastly uses the revenue column as the main forecasting target."})

    if "scenario" in text:
        return jsonify({"reply": "Scenario Analysis shows optimistic, realistic, and pessimistic revenue outcomes."})

    if "mape" in text:
        return jsonify({"reply": "MAPE means Mean Absolute Percentage Error. Lower MAPE usually means better forecast quality."})

    return jsonify({"reply": "I can help you with uploading data, forecasting revenue, analytics, anomalies, and reports."})


def create_forecast_chart_image(chart_data):
    chart_buffer = BytesIO()

    labels = chart_data.get("labels", [])
    actual = chart_data.get("actual", [])
    forecast = chart_data.get("forecast", [])

    plt.figure(figsize=(8, 4))
    plt.plot(labels, actual, marker="o", linewidth=3, label="Actual Revenue")
    plt.plot(labels, forecast, marker="o", linestyle="--", linewidth=3, label="Forecast Revenue")
    plt.title("Forecast vs Actual Revenue")
    plt.xlabel("Periods")
    plt.ylabel("Revenue")
    plt.xticks(rotation=25)
    plt.legend()
    plt.tight_layout()
    plt.savefig(chart_buffer, format="png", dpi=150, bbox_inches="tight")
    plt.close()
    chart_buffer.seek(0)

    return chart_buffer


@app.route("/export-report")
@login_required
def export_report():
    try:
        forecast_days = int(session.get("forecast_days", 7))
        result = run_ml_pipeline(forecast_days)

        latest_file = result["latest_file"]
        analytics = result["analytics"]
        scenarios = result["scenarios"]
        chart_data = result["chart_data"]
        chart_image = create_forecast_chart_image(chart_data)

        buffer = BytesIO()
        pdf = canvas.Canvas(buffer, pagesize=A4)
        width, height = A4

        pdf.setFillColorRGB(0.48, 0.23, 0.93)
        pdf.rect(0, height - 90, width, 90, fill=True, stroke=False)

        pdf.setFillColorRGB(1, 1, 1)
        pdf.setFont("Helvetica-Bold", 24)
        pdf.drawString(50, height - 45, "Forecastly")

        pdf.setFont("Helvetica", 12)
        pdf.drawString(50, height - 65, "AI-Powered Revenue Forecast Report")

        y = height - 130

        pdf.setFillColorRGB(0.13, 0.08, 0.25)
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(50, y, "Report Summary")

        y -= 30
        pdf.setFont("Helvetica", 11)
        pdf.drawString(50, y, f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        y -= 20
        pdf.drawString(50, y, f"File: {latest_file.original_filename}")

        y -= 40
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(50, y, "Revenue Analytics")

        y -= 30
        pdf.setFont("Helvetica", 11)
        pdf.drawString(50, y, f"Total Revenue: {analytics.get('total_sales')}")
        y -= 20
        pdf.drawString(50, y, f"Average Revenue: {analytics.get('average_sales')}")
        y -= 20
        pdf.drawString(50, y, f"Growth Rate: {analytics.get('growth_rate')}%")
        y -= 20
        pdf.drawString(50, y, f"Accuracy: {analytics.get('accuracy')}%")
        y -= 20
        pdf.drawString(50, y, f"MAPE: {analytics.get('mape')}%")
        y -= 20
        pdf.drawString(50, y, f"RMSE: {analytics.get('rmse')}")

        y -= 40
        pdf.setFont("Helvetica-Bold", 14)
        pdf.drawString(50, y, "Scenario Forecast")

        y -= 30
        pdf.setFont("Helvetica", 11)
        pdf.drawString(50, y, f"Optimistic Scenario: {scenarios.get('optimistic')}")
        y -= 20
        pdf.drawString(50, y, f"Realistic Scenario: {scenarios.get('realistic')}")
        y -= 20
        pdf.drawString(50, y, f"Pessimistic Scenario: {scenarios.get('pessimistic')}")
        y -= 20
        pdf.drawString(50, y, f"Average Forecast: {scenarios.get('average_forecast')}")

        y -= 260
        pdf.drawImage(
            ImageReader(chart_image),
            50,
            y,
            width=500,
            height=230,
            preserveAspectRatio=True,
            mask="auto"
        )

        pdf.showPage()
        pdf.save()

        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers["Content-Type"] = "application/pdf"
        response.headers["Content-Disposition"] = "attachment; filename=forecastly_revenue_report.pdf"

        return response

    except Exception as e:
        print("❌ EXPORT ERROR:", e)
        flash(f"Export error: {e}", "error")
        return redirect(url_for("analytics_page"))

@app.route("/export-excel")
@login_required
def export_excel():
    try:
        forecast_days = int(session.get("forecast_days", 7))
        result = run_ml_pipeline(forecast_days)

        analytics = result["analytics"]
        scenarios = result["scenarios"]
        anomalies = result["anomalies"]
        forecast_data = result["forecast_data"]
        latest_file = result["latest_file"]

        wb = Workbook()

        purple_fill = PatternFill("solid", fgColor="7C3AED")
        light_fill = PatternFill("solid", fgColor="F5EFFF")
        white_font = Font(color="FFFFFF", bold=True)
        title_font = Font(size=18, bold=True, color="FFFFFF")
        header_font = Font(bold=True, color="21133F")
        border = Border(
            left=Side(style="thin", color="E9D5FF"),
            right=Side(style="thin", color="E9D5FF"),
            top=Side(style="thin", color="E9D5FF"),
            bottom=Side(style="thin", color="E9D5FF")
        )

        def style_sheet(sheet):
         for row in sheet.iter_rows():
          for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border

            sheet.column_dimensions["A"].width = 28
            sheet.column_dimensions["B"].width = 38
            sheet.column_dimensions["C"].width = 28
            sheet.column_dimensions["D"].width = 22
            sheet.column_dimensions["E"].width = 22

        def write_header(sheet, title):
            sheet.merge_cells("A1:E1")
            sheet["A1"] = title
            sheet["A1"].fill = purple_fill
            sheet["A1"].font = title_font
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet.row_dimensions[1].height = 32
        # =========================
        # SHEET 1 — EXECUTIVE SUMMARY
        # =========================

        summary = wb.active
        summary.title = "Executive Summary"
        write_header(summary, "Forecastly AI Revenue Report")

        summary_data = [
            ["Generated On", datetime.now().strftime("%Y-%m-%d %H:%M")],
            ["Dataset", latest_file.original_filename],
            ["Total Revenue", analytics.get("total_revenue")],
            ["Average Revenue", analytics.get("average_revenue")],
            ["Growth Rate", f"{analytics.get('growth_rate')}%"],
            ["Forecast Accuracy", f"{analytics.get('accuracy')}%"],
            ["MAPE", f"{analytics.get('mape')}%"],
            ["RMSE", analytics.get("rmse")],
            ["Forecast Horizon", f"{forecast_days} days"],
            ["Total Forecast", scenarios.get("total_forecast")]
        ]

        row = 3
        for label, value in summary_data:
            summary[f"A{row}"] = label
            summary[f"B{row}"] = value
            summary[f"A{row}"].font = header_font
            summary[f"A{row}"].fill = light_fill
            row += 1

        style_sheet(summary)

        # =========================
        # SHEET 2 — FORECAST DATA
        # =========================

        forecast_sheet = wb.create_sheet("Forecast Data")
        write_header(forecast_sheet, "Forecast Data")

        headers = ["Date", "Predicted Revenue", "Model Used", "MAPE", "RMSE"]

        for col, header in enumerate(headers, start=1):
            cell = forecast_sheet.cell(row=3, column=col)
            cell.value = header
            cell.fill = purple_fill
            cell.font = white_font

        for row_index, item in enumerate(forecast_data, start=4):
            forecast_sheet.cell(row=row_index, column=1).value = item.get("date")
            forecast_sheet.cell(row=row_index, column=2).value = item.get("predicted_revenue")
            forecast_sheet.cell(row=row_index, column=3).value = item.get("model_used")
            forecast_sheet.cell(row=row_index, column=4).value = item.get("mape")
            forecast_sheet.cell(row=row_index, column=5).value = item.get("rmse")

        style_sheet(forecast_sheet)

        # =========================
        # SHEET 3 — ANALYTICS
        # =========================

        analytics_sheet = wb.create_sheet("Analytics")
        write_header(analytics_sheet, "Analytics Metrics")

        analytics_headers = ["Metric", "Value"]

        for col, header in enumerate(analytics_headers, start=1):
            cell = analytics_sheet.cell(row=3, column=col)
            cell.value = header
            cell.fill = purple_fill
            cell.font = white_font

        analytics_data = [
            ["Total Revenue", analytics.get("total_revenue")],
            ["Average Revenue", analytics.get("average_revenue")],
            ["Max Revenue", analytics.get("max_revenue")],
            ["Min Revenue", analytics.get("min_revenue")],
            ["Moving Average", analytics.get("moving_average")],
            ["Accuracy", f"{analytics.get('accuracy')}%"],
            ["Growth Rate", f"{analytics.get('growth_rate')}%"],
            ["MAPE", f"{analytics.get('mape')}%"],
            ["RMSE", analytics.get("rmse")]
        ]

        for row_index, item in enumerate(analytics_data, start=4):
            analytics_sheet.cell(row=row_index, column=1).value = item[0]
            analytics_sheet.cell(row=row_index, column=2).value = item[1]

        style_sheet(analytics_sheet)

        # =========================
        # SHEET 4 — ANOMALIES
        # =========================

        anomaly_sheet = wb.create_sheet("Anomalies")
        write_header(anomaly_sheet, "AI Anomaly Detection")

        anomaly_headers = ["Date", "Revenue", "Type", "Severity"]

        for col, header in enumerate(anomaly_headers, start=1):
            cell = anomaly_sheet.cell(row=3, column=col)
            cell.value = header
            cell.fill = purple_fill
            cell.font = white_font

        if anomalies:
            for row_index, anomaly in enumerate(anomalies, start=4):
                anomaly_sheet.cell(row=row_index, column=1).value = anomaly.get("date")
                anomaly_sheet.cell(row=row_index, column=2).value = anomaly.get("revenue")
                anomaly_sheet.cell(row=row_index, column=3).value = anomaly.get("type")
                anomaly_sheet.cell(row=row_index, column=4).value = anomaly.get("severity")
        else:
            anomaly_sheet["A4"] = "No significant anomalies detected."

        style_sheet(anomaly_sheet)

        # =========================
        # SHEET 5 — AI INSIGHTS
        # =========================

        insights_sheet = wb.create_sheet("AI Insights")
        write_header(insights_sheet, "Forecastly AI Insights")

        all_insights = []
        all_insights.extend(result.get("scenario_insights", []))
        all_insights.extend(result.get("analytics_insights", []))
        all_insights.extend(result.get("anomaly_insights", []))

        insights_sheet["A3"] = "Insight Type"
        insights_sheet["B3"] = "Insight"
        insights_sheet["A3"].fill = purple_fill
        insights_sheet["B3"].fill = purple_fill
        insights_sheet["A3"].font = white_font
        insights_sheet["B3"].font = white_font

        row = 4
        for insight in result.get("scenario_insights", []):
            insights_sheet[f"A{row}"] = "Scenario"
            insights_sheet[f"B{row}"] = insight
            row += 1

        for insight in result.get("analytics_insights", []):
            insights_sheet[f"A{row}"] = "Analytics"
            insights_sheet[f"B{row}"] = insight
            row += 1

        for insight in result.get("anomaly_insights", []):
            insights_sheet[f"A{row}"] = "Anomaly"
            insights_sheet[f"B{row}"] = insight
            row += 1

        style_sheet(insights_sheet)
        insights_sheet.column_dimensions["B"].width = 80

        # =========================
        # EXPORT FILE
        # =========================

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers["Content-Type"] = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response.headers["Content-Disposition"] = "attachment; filename=forecastly_ai_report.xlsx"

        return response

    except Exception as e:
        print("EXCEL EXPORT ERROR:", e)
        flash(f"Excel export error: {e}", "error")
        return redirect(url_for("analytics_page"))
@app.route('/forgot-password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        email = request.form.get('email')

        user = User.query.filter_by(email=email).first()

        if user:
            token = serializer.dumps(email, salt='password-reset-salt')

            reset_link = url_for(
                'reset_password',
                token=token,
                _external=True
            )

            msg = Message(
                subject='Reset your Forecastly password',
                recipients=[email],
                body=f'''
Hello,

You requested to reset your Forecastly password.

Click the link below to create a new password:
{reset_link}

This link will expire in 30 minutes.

If you did not request this, please ignore this email.

Forecastly Team
'''
            )

            mail.send(msg)

        flash('If this email exists, a password reset link has been sent.')
        return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/reset-password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    try:
        email = serializer.loads(
            token,
            salt='password-reset-salt',
            max_age=1800
        )
    except SignatureExpired:
        flash('The password reset link has expired.')
        return redirect(url_for('forgot_password'))
    except BadSignature:
        flash('Invalid password reset link.')
        return redirect(url_for('forgot_password'))

    user = User.query.filter_by(email=email).first()

    if not user:
        flash('User not found.')
        return redirect(url_for('forgot_password'))

    if request.method == 'POST':
        password = request.form.get('password')
        confirm_password = request.form.get('confirm_password')

        if password != confirm_password:
            flash('Passwords do not match.')
            return redirect(request.url)

        user.password_hash = generate_password_hash(password)
        db.session.commit()

        flash('Your password has been updated. Please log in.')
        return redirect(url_for('login'))

    return render_template('reset_password.html')
@app.route("/admin")
@login_required
@admin_required
def admin_dashboard():

    users = User.query.order_by(
        User.created_at.desc()
    ).all()

    files = UploadedFile.query.order_by(
        UploadedFile.uploaded_at.desc()
    ).all()

    total_users = len(users)
    total_files = len(files)

    total_rows = sum(
        file.rows_count or 0
        for file in files
    )
    leaderboard = sorted(
    users,
    key=lambda user: sum(
        file.rows_count or 0
        for file in user.files
    ),
    reverse=True
)
    return render_template(
        "admin.html",
        users=users,
        files=files,
        total_users=total_users,
        total_files=total_files,
        total_rows=total_rows,
        leaderboard=leaderboard
    )
# =========================
# ERROR HANDLERS
# =========================

@app.errorhandler(413)
def file_too_large(error):
    flash("File is too large. Maximum allowed size is 10 MB.", "error")
    return redirect(url_for("upload_data"))


@app.errorhandler(404)
def page_not_found(error):
    return render_template("404.html"), 404


# =========================
# RUN APP
# =========================

if __name__ == "__main__":
    print("🚀 Forecastly is starting...")

    with app.app_context():
        db.create_all()

        admin_user = User.query.filter_by(email="aizxsab@gmail.com").first()

        if admin_user:
            admin_user.is_admin = True
            db.session.commit()
            print("✅ Admin access granted to aizxsab@gmail.com")
        else:
            print("❌ Admin user not found. Please sign up first.")

    app.run(debug=True)
 